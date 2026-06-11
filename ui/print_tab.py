"""
打印录入页面 UI 组件
支持自由选择 BarTender 模板和 Excel 校验文件
自动保存配置到本地文件
"""

import os
import re
import json
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFormLayout, 
    QLabel, QLineEdit, QSpinBox, QPushButton, QComboBox,
    QMessageBox, QGroupBox, QTextEdit, QFileDialog,
    QDialog, QDialogButtonBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QFrame, QCheckBox
)
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont


class PrintTab(QWidget):
    """打印录入标签页"""
    
    print_success = pyqtSignal(str)
    
    def __init__(self, record_manager, default_template_path: str, default_excel_path: str = ""):
        super().__init__()
        
        self.record_manager = record_manager
        self.config_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config.json")
        self.user_config = self.load_user_config()
        
        # 使用用户保存的路径，如果没有则使用默认路径
        self.template_path = self.user_config.get("template_path", default_template_path)
        self.validation_excel_path = self.user_config.get("excel_path", default_excel_path)
        self.bartender = None
        self.data_sources = []
        self.validation_data = set()
        self.auto_skip_duplicate = self.user_config.get("auto_skip_duplicate", False)
        
        self.init_ui()
        self.load_template()
        if self.validation_excel_path:
            self.load_excel_data(silent=True)
    
    def load_user_config(self):
        """加载用户配置"""
        config = {
            "template_path": "",
            "excel_path": "",
            "auto_skip_duplicate": False
        }
        
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    saved_config = json.load(f)
                    config.update(saved_config)
            except Exception as e:
                print(f"加载配置文件失败：{e}")
        
        return config
    
    def save_user_config(self):
        """保存用户配置"""
        config = {
            "template_path": self.template_path,
            "excel_path": self.validation_excel_path,
            "auto_skip_duplicate": self.auto_skip_duplicate
        }
        
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"保存配置文件失败：{e}")
            return False
    
    def init_ui(self):
        """初始化 UI"""
        main_layout = QVBoxLayout()
        main_layout.setSpacing(15)
        
        if self.layout():
            while self.layout().count():
                item = self.layout().takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
        else:
            self.setLayout(main_layout)
            main_layout = self.layout()
        
        main_layout.setAlignment(Qt.AlignTop)
        
        # =====================================================
        # 1. 模板文件选择区域
        # =====================================================
        template_group = QGroupBox("BarTender 模板配置")
        template_layout = QHBoxLayout()
        template_layout.setSpacing(10)
        
        self.template_path_label = QLabel("未选择模板")
        self.template_path_label.setStyleSheet("color: #999;")
        self.template_path_label.setFixedWidth(300)
        
        self.select_template_btn = QPushButton("选择模板")
        self.select_template_btn.clicked.connect(self.on_select_template_clicked)
        
        self.template_status = QLabel("未加载")
        self.template_status.setStyleSheet("color: #f44336; font-weight: bold;")
        
        template_layout.addWidget(QLabel("模板:"))
        template_layout.addWidget(self.template_path_label)
        template_layout.addWidget(self.select_template_btn)
        template_layout.addWidget(QLabel("状态:"))
        template_layout.addWidget(self.template_status)
        
        template_group.setLayout(template_layout)
        
        # =====================================================
        # 2. Excel 文件选择区域
        # =====================================================
        excel_group = QGroupBox("校验数据源配置")
        excel_layout = QHBoxLayout()
        excel_layout.setSpacing(10)
        
        self.excel_path_label = QLabel("未选择文件")
        self.excel_path_label.setStyleSheet("color: #999;")
        self.excel_path_label.setFixedWidth(300)
        
        self.select_excel_btn = QPushButton("选择 Excel")
        self.select_excel_btn.clicked.connect(self.on_select_excel_clicked)
        
        self.load_excel_btn = QPushButton("加载数据")
        self.load_excel_btn.clicked.connect(self.on_load_excel_clicked)
        self.load_excel_btn.setEnabled(False)
        
        self.validation_status = QLabel("未校验")
        self.validation_status.setStyleSheet("color: #f44336; font-weight: bold;")
        
        excel_layout.addWidget(QLabel("Excel:"))
        excel_layout.addWidget(self.excel_path_label)
        excel_layout.addWidget(self.select_excel_btn)
        excel_layout.addWidget(self.load_excel_btn)
        excel_layout.addWidget(QLabel("状态:"))
        excel_layout.addWidget(self.validation_status)
        
        excel_group.setLayout(excel_layout)
        
        # =====================================================
        # 3. 分隔线
        # =====================================================
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        line.setStyleSheet("background: #d0d0d0;")
        line.setFixedHeight(2)
        
        # =====================================================
        # 4. 操作按钮区域
        # =====================================================
        button_layout = QHBoxLayout()
        button_layout.setSpacing(20)
        
        self.print_btn = QPushButton("打印标签")
        self.print_btn.setMinimumHeight(60)
        self.print_btn.setFont(QFont("Microsoft YaHei", 16, QFont.Bold))
        self.print_btn.clicked.connect(self.on_print_clicked)
        self.print_btn.setEnabled(False)
        
        self.preview_btn = QPushButton("查看数据")
        self.preview_btn.setMinimumHeight(60)
        self.preview_btn.clicked.connect(self.on_preview_excel_clicked)
        self.preview_btn.setEnabled(False)
        
        button_layout.addWidget(self.print_btn)
        button_layout.addWidget(self.preview_btn)
        
        # =====================================================
        # 5. 高级选项
        # =====================================================
        advanced_group = QGroupBox("高级选项")
        advanced_layout = QHBoxLayout()
        
        self.skip_duplicate_check = QCheckBox("自动跳过重复打印的 IMEI（不弹窗提示）")
        self.skip_duplicate_check.setChecked(self.auto_skip_duplicate)
        self.skip_duplicate_check.stateChanged.connect(self.on_skip_duplicate_changed)
        
        advanced_layout.addWidget(self.skip_duplicate_check)
        advanced_layout.addStretch()
        
        advanced_group.setLayout(advanced_layout)
        
        # =====================================================
        # 6. 状态显示区域
        # =====================================================
        status_group = QGroupBox("打印状态")
        status_layout = QVBoxLayout()
        
        self.status_label = QLabel("就绪 - 请先选择模板和 Excel 文件")
        self.status_label.setFont(QFont("Microsoft YaHei", 12))
        self.status_label.setStyleSheet("color: #FF9800; font-weight: bold;")
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(200)
        self.log_text.setPlaceholderText("打印日志...")
        
        status_layout.addWidget(self.status_label)
        status_layout.addWidget(self.log_text)
        status_group.setLayout(status_layout)
        
        # 添加到主布局
        main_layout.addWidget(template_group)
        main_layout.addWidget(excel_group)
        main_layout.addWidget(line)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(advanced_group)
        main_layout.addWidget(status_group)
    
    def load_template(self, silent=False):
        """加载模板文件"""
        if not self.template_path:
            self.template_status.setText("未选择")
            self.template_status.setStyleSheet("color: #f44336; font-weight: bold;")
            if not silent:
                self.log_message("模板未加载")
            return False
        
        if not os.path.exists(self.template_path):
            self.template_status.setText("文件不存在")
            self.template_status.setStyleSheet("color: #f44336; font-weight: bold;")
            error_msg = (
                f"模板文件不存在或无法访问！\n\n"
                f"文件路径：{self.template_path}\n\n"
                f"可能原因：\n"
                f"1. 文件已被移动或删除\n"
                f"2. 路径中有中文字符或特殊字符\n"
                f"3. 没有访问该文件的权限\n"
                f"4. 网络驱动器未连接\n\n"
                f"请点击「选择模板」重新选择文件。"
            )
            if not silent:
                QMessageBox.critical(self, "模板加载失败", error_msg)
            self.log_message(f"模板文件不存在：{self.template_path}")
            return False
        
        try:
            from core.bartender import BarTenderPrinter
            
            if not silent:
                self.template_status.setText("加载中...")
                self.template_status.setStyleSheet("color: #2196F3; font-weight: bold;")
            
            self.bartender = BarTenderPrinter(self.template_path)
            success, result = self.bartender.load_template()
            
            if success:
                success, data_sources = self.bartender.get_data_sources()
                if success and data_sources:
                    self.data_sources = data_sources
                    if not silent:
                        self.template_status.setText(f"已加载 {len(data_sources)} 个数据源")
                        self.template_status.setStyleSheet("color: #4CAF50; font-weight: bold;")
                        self.log_message(f"模板加载成功：{os.path.basename(self.template_path)}")
                else:
                    if not silent:
                        self.template_status.setText("未找到数据源")
                        self.template_status.setStyleSheet("color: #FF9800; font-weight: bold;")
                        self.log_message("未找到数据源，使用默认配置")
                self.check_ready_status()
                return True
            else:
                self.template_status.setText("加载失败")
                self.template_status.setStyleSheet("color: #f44336; font-weight: bold;")
                error_msg = (
                    f"无法加载 BarTender 模板！\n\n"
                    f"模板文件：{self.template_path}\n\n"
                    f"错误信息：{result}\n\n"
                    f"可能原因：\n"
                    f"1. BarTender 未安装或未正确激活\n"
                    f"2. 模板文件损坏或格式错误\n"
                    f"3. BarTender 版本不兼容\n"
                    f"4. COM 组件初始化失败\n\n"
                    f"请检查：\n"
                    f"- BarTender 2021 是否已安装并激活\n"
                    f"- 模板文件是否可以正常打开\n"
                    f"- 尝试在 BarTender 中打开此模板"
                )
                if not silent:
                    QMessageBox.critical(self, "模板加载失败", error_msg)
                self.log_message(f"模板加载失败：{result}")
                return False
                
        except Exception as e:
            self.template_status.setText("加载失败")
            self.template_status.setStyleSheet("color: #f44336; font-weight: bold;")
            error_msg = (
                f"加载模板时发生错误！\n\n"
                f"模板文件：{self.template_path}\n\n"
                f"错误详情：{str(e)}\n\n"
                f"可能原因：\n"
                f"1. BarTender 服务未启动\n"
                f"2. 系统权限问题\n"
                f"3. 内存不足\n\n"
                f"请尝试：\n"
                f"- 重启 BarTender 软件\n"
                f"- 以管理员身份运行本程序\n"
                f"- 重新启动电脑"
            )
            if not silent:
                QMessageBox.critical(self, "错误", error_msg)
            self.log_message(f"初始化失败：{str(e)}")
            return False
    
    def check_ready_status(self):
        """检查是否可以打印"""
        can_print = (
            self.template_path and 
            os.path.exists(self.template_path) and
            self.validation_data and
            len(self.validation_data) > 0
        )
        
        if can_print:
            self.print_btn.setEnabled(True)
            self.preview_btn.setEnabled(True)
            self.status_label.setText("就绪 - 可以点击打印")
            self.status_label.setStyleSheet("color: #4CAF50; font-weight: bold;")
        else:
            self.print_btn.setEnabled(False)
            self.preview_btn.setEnabled(False)
    
    def log_message(self, message: str):
        """记录日志"""
        from datetime import datetime
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{timestamp}] {message}")
    
    def update_status(self, status: str, success: bool = True):
        """更新状态显示"""
        self.status_label.setText(status)
        if success:
            self.status_label.setStyleSheet("color: #4CAF50; font-weight: bold;")
        else:
            self.status_label.setStyleSheet("color: #f44336; font-weight: bold;")
    
    def on_skip_duplicate_changed(self, state):
        """跳过重复打印选项改变"""
        self.auto_skip_duplicate = (state == Qt.Checked)
        self.save_user_config()
    
    def on_select_template_clicked(self):
        """选择 BarTender 模板文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择 BarTender 模板文件",
            "",
            "BarTender 模板 (*.btw)"
        )
        
        if file_path:
            self.template_path = file_path
            self.template_path_label.setText(os.path.basename(file_path))
            self.template_path_label.setToolTip(file_path)
            self.template_path_label.setStyleSheet("color: #2196F3; font-weight: bold;")
            self.template_status.setText("待加载")
            self.template_status.setStyleSheet("color: #2196F3; font-weight: bold;")
            self.data_sources = []
            self.log_message(f"已选择模板：{file_path}")
            self.save_user_config()
            
            if self.load_template():
                self.log_message("模板加载成功，配置已保存")
    
    def on_select_excel_clicked(self):
        """选择 Excel 文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择校验 Excel 文件",
            "",
            "Excel 文件 (*.xlsx *.xls)"
        )
        
        if file_path:
            self.validation_excel_path = file_path
            self.excel_path_label.setText(os.path.basename(file_path))
            self.excel_path_label.setToolTip(file_path)
            self.excel_path_label.setStyleSheet("color: #2196F3; font-weight: bold;")
            self.load_excel_btn.setEnabled(True)
            self.validation_data.clear()
            self.validation_status.setText("待加载")
            self.validation_status.setStyleSheet("color: #FF9800; font-weight: bold;")
            self.print_btn.setEnabled(False)
            self.preview_btn.setEnabled(False)
            self.update_status("已选择 Excel 文件，请点击加载数据", True)
            self.log_message(f"已选择校验文件：{file_path}")
            self.save_user_config()
    
    def on_load_excel_clicked(self):
        """加载 Excel 数据"""
        self.load_excel_data(silent=False)
    
    def load_excel_data(self, silent=False):
        """加载 Excel 数据的内部方法"""
        if not self.validation_excel_path:
            if not silent:
                QMessageBox.warning(self, "错误", "请先选择 Excel 文件")
            return False
        
        if not os.path.exists(self.validation_excel_path):
            error_msg = (
                f"Excel 文件不存在或无法访问！\n\n"
                f"文件路径：{self.validation_excel_path}\n\n"
                f"可能原因：\n"
                f"1. 文件已被移动或删除\n"
                f"2. 文件正在被其他程序占用\n"
                f"3. 没有访问该文件的权限\n"
                f"4. 网络驱动器未连接\n\n"
                f"请点击「选择 Excel」重新选择文件。"
            )
            if not silent:
                QMessageBox.critical(self, "文件错误", error_msg)
            self.log_message(f"Excel 文件不存在：{self.validation_excel_path}")
            return False
        
        try:
            if not silent:
                self.validation_status.setText("加载中...")
                self.validation_status.setStyleSheet("color: #2196F3; font-weight: bold;")
                self.update_status("正在加载 Excel 数据...", True)
            
            from openpyxl import load_workbook
            wb = load_workbook(self.validation_excel_path, read_only=True)
            ws = wb.active
            
            self.validation_data.clear()
            count = 0
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        value = str(cell.value).strip()
                        self.validation_data.add(value)
                        count += 1
            
            wb.close()
            
            if count > 0:
                if not silent:
                    self.validation_status.setText(f"已加载 {count} 条")
                    self.validation_status.setStyleSheet("color: #4CAF50; font-weight: bold;")
                    self.check_ready_status()
                    self.log_message(f"成功加载 {count} 条数据")
                    QMessageBox.information(
                        self,
                        "加载成功",
                        f"成功加载 {count} 条数据\n\n"
                        f"只有完全匹配的数据才能打印"
                    )
                return True
            else:
                self.validation_status.setText("无数据")
                self.validation_status.setStyleSheet("color: #f44336; font-weight: bold;")
                if not silent:
                    self.update_status("Excel 文件中没有数据", False)
                    self.log_message("Excel 文件中没有数据")
                    QMessageBox.warning(
                        self,
                        "警告",
                        "Excel 文件中没有数据\n\n"
                        "请确保文件中包含有效数据"
                    )
                return False
                
        except PermissionError:
            self.validation_status.setText("权限错误")
            self.validation_status.setStyleSheet("color: #f44336; font-weight: bold;")
            error_msg = (
                f"无法访问 Excel 文件！\n\n"
                f"文件路径：{self.validation_excel_path}\n\n"
                f"可能原因：\n"
                f"1. 文件正在被 Excel 或其他程序占用\n"
                f"2. 没有读取该文件的权限\n"
                f"3. 文件被杀毒软件阻止\n\n"
                f"请尝试：\n"
                f"- 关闭 Excel 中打开的该文件\n"
                f"- 以管理员身份运行本程序\n"
                f"- 检查文件权限"
            )
            if not silent:
                self.update_status("权限错误 - 文件可能被占用", False)
                self.log_message(f"Excel 文件权限错误：{self.validation_excel_path}")
                QMessageBox.critical(self, "权限错误", error_msg)
            return False
            
        except Exception as e:
            self.validation_status.setText("加载失败")
            self.validation_status.setStyleSheet("color: #f44336; font-weight: bold;")
            error_msg = (
                f"无法加载 Excel 文件！\n\n"
                f"文件路径：{self.validation_excel_path}\n\n"
                f"错误详情：{str(e)}\n\n"
                f"可能原因：\n"
                f"1. 文件格式不支持（需要 .xlsx 或.xls）\n"
                f"2. 文件已损坏\n"
                f"3. 文件是加密的或有密码保护\n\n"
                f"请检查：\n"
                f"- 文件格式是否正确\n"
                f"- 尝试用 Excel 打开此文件\n"
                f"- 确保文件没有密码保护"
            )
            if not silent:
                self.update_status(f"加载失败：{str(e)}", False)
                self.log_message(f"加载 Excel 失败：{str(e)}")
                QMessageBox.critical(self, "加载失败", error_msg)
            return False
    
    def on_preview_excel_clicked(self):
        """预览 Excel 数据"""
        if not self.validation_data:
            QMessageBox.information(self, "无数据", "没有可预览的数据")
            return
        
        dialog = ExcelPreviewDialog(self.validation_data, self)
        dialog.exec_()
    
    def on_print_clicked(self):
        """打印按钮点击事件"""
        # 弹出输入对话框（支持回车打印）
        dialog = IMEIInputDialog(self)
        if dialog.exec_() != QDialog.Accepted:
            return
        
        imei = dialog.get_imei().strip()
        
        if not imei:
            QMessageBox.warning(self, "验证失败", "请输入 IMEI 号码")
            return
        
        # 校验是否在 Excel 中
        if self.validation_data and imei not in self.validation_data:
            QMessageBox.critical(
                self,
                "校验失败",
                f"IMEI {imei} 不在校验数据中！\n\n"
                f"请确保输入的数据与 Excel 文件中的数据完全匹配"
            )
            self.log_message(f"校验失败：IMEI {imei} 不在允许列表中")
            self.update_status("校验失败 - IMEI 不在允许列表中", False)
            return
        
        copies = dialog.get_copies()
        
        # 获取数据源名称 - 如果只有一个数据源，自动使用
        if len(self.data_sources) == 1:
            data_source = self.data_sources[0]
        elif len(self.data_sources) == 0:
            data_source = "IMEI1"  # 默认值
        else:
            # 多个数据源时需要用户选择
            dialog_ds = DataSourceDialog(self.data_sources, self)
            if dialog_ds.exec_() != QDialog.Accepted:
                return
            data_source = dialog_ds.get_data_source()
        
        import getpass
        operator = getpass.getuser()
        
        # 检查是否已打印
        exists, record = self.record_manager.check_imei_exists(imei)
        
        if exists:
            if self.auto_skip_duplicate:
                self.log_message(f"自动跳过重复打印：IMEI {imei}")
                self.update_status("已跳过重复 IMEI", False)
                return
            else:
                dialog_dup = DuplicatePrintDialog(imei, record, self)
                result = dialog_dup.exec_()
                
                if result == QDialog.Rejected:
                    self.update_status("已取消打印", False)
                    self.log_message(f"用户取消打印已存在的 IMEI: {imei}")
                    return
        
        # 执行打印
        self.update_status("正在打印...", True)
        self.print_btn.setEnabled(False)
        
        try:
            if not self.bartender:
                self.load_template()
            
            success, msg = self.bartender.set_data_value(data_source, imei)
            if not success:
                raise Exception(msg)
            
            success, msg = self.bartender.print_label(copies)
            if not success:
                raise Exception(msg)
            
            self.record_manager.add_record(imei, copies, operator)
            
            self.log_message(f"打印成功：IMEI={imei}, 份数={copies}, 数据源={data_source}")
            self.update_status(f"打印成功！已打印 {copies} 份", True)
            self.print_success.emit(imei)
            
            QMessageBox.information(self, "打印成功", f"标签已成功打印 {copies} 份！")
            
        except Exception as e:
            error_msg = str(e)
            self.log_message(f"打印失败：{error_msg}")
            self.update_status(f"打印失败：{error_msg}", False)
            QMessageBox.critical(self, "打印失败", f"打印过程中发生错误:\n{error_msg}")
        finally:
            self.print_btn.setEnabled(True)


class IMEIInputDialog(QDialog):
    """IMEI 输入对话框 - 支持回车快速打印"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        
        self.setWindowTitle("输入 IMEI 号码")
        self.setModal(True)
        self.resize(450, 280)
        self.init_ui()
    
    def init_ui(self):
        """初始化 UI"""
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        # IMEI 输入
        imei_group = QGroupBox("IMEI 信息")
        imei_form = QFormLayout()
        imei_form.setSpacing(10)
        
        self.imei_input = QLineEdit()
        self.imei_input.setPlaceholderText("请输入或扫描 IMEI")
        self.imei_input.setFont(QFont("Consolas", 18))
        self.imei_input.returnPressed.connect(self.on_print)  # 回车直接打印
        
        self.copies_spin = QSpinBox()
        self.copies_spin.setMinimum(1)
        self.copies_spin.setMaximum(999)
        self.copies_spin.setValue(1)
        self.copies_spin.setFont(QFont("Microsoft YaHei", 14))
        
        imei_form.addRow("IMEI 号码:", self.imei_input)
        imei_form.addRow("打印份数:", self.copies_spin)
        
        imei_group.setLayout(imei_form)
        
        # 提示信息
        info_label = QLabel("ℹ 只有 Excel 文件中完全匹配的 IMEI 才能打印")
        info_label.setStyleSheet("color: #2196F3; font-size: 12px;")
        info_label.setWordWrap(True)
        
        # 按钮
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        button_box.button(QDialogButtonBox.Ok).setText("打印")
        button_box.button(QDialogButtonBox.Cancel).setText("取消")
        
        layout.addWidget(imei_group)
        layout.addWidget(info_label)
        layout.addWidget(button_box)
        
        self.setLayout(layout)
        self.imei_input.setFocus()
    
    def on_print(self):
        """回车键触发打印"""
        imei = self.imei_input.text().strip()
        if imei:
            self.accept()
        else:
            QMessageBox.warning(self, "输入错误", "请输入 IMEI")
    
    def get_imei(self) -> str:
        """获取 IMEI"""
        return self.imei_input.text()
    
    def get_copies(self) -> int:
        """获取打印份数"""
        return self.copies_spin.value()


class DataSourceDialog(QDialog):
    """数据源选择对话框"""
    
    def __init__(self, data_sources: list, parent=None):
        super().__init__(parent)
        
        self.data_sources = data_sources
        self.setWindowTitle("选择数据源")
        self.setModal(True)
        self.resize(400, 300)
        self.init_ui()
    
    def init_ui(self):
        """初始化 UI"""
        layout = QVBoxLayout()
        layout.setSpacing(15)
        
        # 标题
        title_label = QLabel("模板包含多个数据源，请选择要使用的数据源：")
        title_label.setWordWrap(True)
        layout.addWidget(title_label)
        
        # 数据源列表
        self.combo = QComboBox()
        self.combo.setFont(QFont("Microsoft YaHei", 14))
        for ds in self.data_sources:
            self.combo.addItem(ds)
        
        layout.addWidget(QLabel("数据源:"))
        layout.addWidget(self.combo)
        
        # 按钮
        button_box = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        button_box.button(QDialogButtonBox.Ok).setText("确定")
        button_box.button(QDialogButtonBox.Cancel).setText("取消")
        
        layout.addWidget(button_box)
        self.setLayout(layout)
    
    def get_data_source(self) -> str:
        """获取选中的数据源"""
        return self.combo.currentText()


class ExcelPreviewDialog(QDialog):
    """Excel 数据预览对话框"""
    
    def __init__(self, data: set, parent=None):
        super().__init__(parent)
        
        self.data = data
        
        self.setWindowTitle("Excel 数据预览")
        self.setModal(True)
        self.resize(600, 500)
        self.init_ui()
    
    def init_ui(self):
        """初始化 UI"""
        layout = QVBoxLayout()
        
        # 统计信息
        info_label = QLabel(f"共 {len(self.data)} 条数据")
        info_label.setFont(QFont("Microsoft YaHei", 12, QFont.Bold))
        info_label.setStyleSheet("color: #2196F3;")
        
        # 数据表格
        self.table = QTableWidget()
        self.table.setColumnCount(1)
        self.table.setHorizontalHeaderLabels(["数据内容"])
        
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        
        self.table.setAlternatingRowColors(True)
        
        # 加载数据（限制显示前 100 条）
        display_data = sorted(list(self.data))[:100]
        self.table.setRowCount(len(display_data))
        
        for i, item in enumerate(display_data):
            self.table.setItem(i, 0, QTableWidgetItem(item))
        
        # 提示
        if len(self.data) > 100:
            tip_label = QLabel(f"（仅显示前 100 条，共 {len(self.data)} 条）")
            tip_label.setStyleSheet("color: #999; font-size: 11px;")
            layout.addWidget(tip_label)
        
        # 关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.accept)
        
        layout.addWidget(info_label)
        layout.addWidget(self.table)
        layout.addWidget(close_btn)
        
        self.setLayout(layout)


class DuplicatePrintDialog(QMessageBox):
    """重复打印确认对话框"""
    
    def __init__(self, imei: str, record: dict, parent=None):
        super().__init__(parent)
        
        self.setWindowTitle("重复打印警告")
        self.setIcon(QMessageBox.Warning)
        
        print_time = record.get("print_time", "未知")
        copies = record.get("copies", 0)
        operator = record.get("operator", "未知")
        
        message = (
            f"IMEI {imei} 已打印过！\n\n"
            f"上次打印时间：{print_time}\n"
            f"上次打印份数：{copies}\n"
            f"操作员：{operator}\n\n"
            f"是否继续打印？"
        )
        
        self.setText(message)
        self.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        self.button(QMessageBox.Yes).setText("继续打印")
        self.button(QMessageBox.No).setText("放弃打印")
        self.setDefaultButton(QMessageBox.No)


from PyQt5.QtWidgets import QDialog


if __name__ == "__main__":
    from PyQt5.QtWidgets import QApplication
    import sys
    
    app = QApplication(sys.argv)
    from core.excel_manager import ExcelRecordManager
    
    record_manager = ExcelRecordManager("data/records.xlsx")
    template = PrintTab(record_manager, "templates/sample.btw")
    template.show()
    sys.exit(app.exec_())
