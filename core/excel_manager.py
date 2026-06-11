"""
Excel 数据记录管理模块
管理已打印 IMEI 记录
"""

import os
import sqlite3
from datetime import datetime
from typing import List, Dict, Optional, Tuple


class ExcelRecordManager:
    """打印记录管理器"""
    
    def __init__(self, excel_path: str):
        """
        初始化记录管理器
        
        Args:
            excel_path: Excel 文件路径
        """
        self.excel_path = os.path.abspath(excel_path)
        self._ensure_file_exists()
    
    def _ensure_file_exists(self):
        """确保 Excel 文件存在，不存在则创建"""
        if not os.path.exists(self.excel_path):
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "打印记录"
                ws.append(["IMEI", "打印时间", "份数", "操作员", "备注"])
                wb.save(self.excel_path)
            except Exception as e:
                print(f"创建 Excel 文件失败：{e}")
    
    def check_imei_exists(self, imei: str) -> Tuple[bool, Optional[Dict]]:
        """
        检查 IMEI 是否已打印
        
        Args:
            imei: IMEI 号码
            
        Returns:
            (exists: bool, record: dict or None)
        """
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and str(row[0]).strip() == imei.strip():
                    return True, {
                        "imei": row[0],
                        "print_time": row[1],
                        "copies": row[2],
                        "operator": row[3],
                        "note": row[4]
                    }
            
            return False, None
        except Exception as e:
            print(f"检查 IMEI 失败：{e}")
            return False, None
    
    def add_record(self, imei: str, copies: int, operator: str = "admin", note: str = "") -> Tuple[bool, str]:
        """
        添加打印记录
        
        Args:
            imei: IMEI 号码
            copies: 打印份数
            operator: 操作员
            note: 备注
            
        Returns:
            (success: bool, message: str)
        """
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            print_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([imei, print_time, copies, operator, note])
            
            wb.save(self.excel_path)
            return True, "记录添加成功"
        except Exception as e:
            return False, f"添加记录失败：{str(e)}"
    
    def get_records(self, start_date: Optional[str] = None, end_date: Optional[str] = None, 
                    imei_keyword: Optional[str] = None) -> List[Dict]:
        """
        获取打印记录
        
        Args:
            start_date: 开始日期 (YYYY-MM-DD)
            end_date: 结束日期 (YYYY-MM-DD)
            imei_keyword: IMEI 关键字
            
        Returns:
            记录列表
        """
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                record = {
                    "imei": row[0] or "",
                    "print_time": row[1] or "",
                    "copies": row[2] or 0,
                    "operator": row[3] or "",
                    "note": row[4] or ""
                }
                
                if start_date and record["print_time"] < start_date:
                    continue
                if end_date and record["print_time"] > end_date + " 23:59:59":
                    continue
                if imei_keyword and imei_keyword not in record["imei"]:
                    continue
                
                records.append(record)
            
            return records
        except Exception as e:
            print(f"获取记录失败：{e}")
            return []
    
    def get_statistics(self, group_by: str = "day") -> List[Dict]:
        """
        获取统计数据
        
        Args:
            group_by: 分组方式 (day/week/month/operator)
            
        Returns:
            统计数据列表
        """
        try:
            from openpyxl import load_workbook
            import pandas as pd
            
            df = pd.DataFrame(self.get_records())
            if df.empty:
                return []
            
            df["print_time"] = pd.to_datetime(df["print_time"])
            
            if group_by == "day":
                df["date"] = df["print_time"].dt.strftime("%Y-%m-%d")
                grouped = df.groupby("date").agg({
                    "imei": "count",
                    "copies": "sum"
                }).reset_index()
                grouped.columns = ["date", "count", "total_copies"]
            elif group_by == "month":
                df["date"] = df["print_time"].dt.strftime("%Y-%m")
                grouped = df.groupby("date").agg({
                    "imei": "count",
                    "copies": "sum"
                }).reset_index()
                grouped.columns = ["date", "count", "total_copies"]
            elif group_by == "operator":
                grouped = df.groupby("operator").agg({
                    "imei": "count",
                    "copies": "sum"
                }).reset_index()
                grouped.columns = ["operator", "count", "total_copies"]
            else:
                return []
            
            return grouped.to_dict("records")
        except Exception as e:
            print(f"获取统计数据失败：{e}")
            return []
    
    def export_to_excel(self, output_path: str, records: List[Dict]) -> Tuple[bool, str]:
        """
        导出记录到 Excel
        
        Args:
            output_path: 输出路径
            records: 记录列表
            
        Returns:
            (success: bool, message: str)
        """
        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "导出记录"
            
            ws.append(["IMEI", "打印时间", "份数", "操作员", "备注"])
            
            for record in records:
                ws.append([
                    record.get("imei", ""),
                    record.get("print_time", ""),
                    record.get("copies", 0),
                    record.get("operator", ""),
                    record.get("note", "")
                ])
            
            wb.save(output_path)
            return True, f"导出成功：{output_path}"
        except Exception as e:
            return False, f"导出失败：{str(e)}"
